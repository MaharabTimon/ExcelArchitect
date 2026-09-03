import os
import sys
import pandas as pd
import numpy as np
import traceback
import re
import io
import openpyxl
import subprocess
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pathlib import Path
from contextlib import redirect_stdout

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTextEdit, QLabel, QFileDialog, QInputDialog,
    QFrame, QSplitter, QLineEdit, QCheckBox, QGraphicsDropShadowEffect,
    QComboBox, QMessageBox
)
from PyQt6.QtCore import Qt, pyqtSignal, QObject, QThread, QTimer
from PyQt6.QtGui import QColor, QTextCursor

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference, LineChart

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
STYLE = """
QMainWindow, QWidget {
    background-color: #0f1117;
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
    color: #e2e8f0;
}
QFrame#card {
    background-color: #1a1d27;
    border: 1px solid #2d3148;
    border-radius: 16px;
}
QPushButton {
    background-color: #1e2235;
    border: 1px solid #2d3148;
    color: #94a3b8;
    padding: 10px 20px;
    border-radius: 10px;
    font-size: 13px;
    font-weight: 600;
}
QPushButton:hover {
    background-color: #252a3d;
    border-color: #4a5380;
    color: #e2e8f0;
}
QPushButton#actionButton {
    background-color: #4f46e5;
    color: #ffffff;
    border: none;
    padding: 12px 24px;
    font-size: 14px;
}
QPushButton#actionButton:hover {
    background-color: #4338ca;
}
QPushButton#actionButton:disabled {
    background-color: #312e81;
    color: #818cf8;
}
QPushButton#undoBtn {
    background-color: #1e2235;
    border: 1px solid #92400e;
    color: #fbbf24;
}
QPushButton#undoBtn:hover {
    background-color: #2a1f0e;
}
QPushButton#undoBtn:disabled {
    background-color: #1e2235;
    border-color: #2d3148;
    color: #4b5563;
}
QTextEdit, QLineEdit {
    background-color: #13161f;
    border: 1px solid #2d3148;
    color: #e2e8f0;
    padding: 14px;
    border-radius: 10px;
    font-size: 14px;
    selection-background-color: #4f46e5;
}
QTextEdit:focus, QLineEdit:focus {
    border: 1px solid #4f46e5;
    background-color: #0f1117;
}
QComboBox {
    background-color: #13161f;
    border: 1px solid #2d3148;
    color: #94a3b8;
    padding: 7px 14px;
    border-radius: 8px;
    font-size: 13px;
}
QComboBox::drop-down { border: none; }
QComboBox QAbstractItemView {
    background-color: #1a1d27;
    border: 1px solid #2d3148;
    color: #e2e8f0;
    selection-background-color: #4f46e5;
}
QCheckBox {
    font-size: 13px;
    color: #64748b;
    spacing: 8px;
}
QCheckBox:hover { color: #94a3b8; }
QCheckBox::indicator {
    width: 16px; height: 16px;
    border-radius: 4px;
    border: 1px solid #2d3148;
    background: #13161f;
}
QCheckBox::indicator:checked {
    background: #4f46e5;
    border-color: #4f46e5;
}
QLabel#header {
    font-size: 22px;
    font-weight: 800;
    color: #e2e8f0;
    letter-spacing: -0.5px;
}
QLabel#badge {
    background-color: #1e1b4b;
    color: #818cf8;
    padding: 3px 10px;
    border-radius: 5px;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1px;
}
QSplitter::handle {
    background-color: #2d3148;
    width: 1px;
}
"""

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def drop_shadow(widget, blur=24, alpha=60, dy=6):
    fx = QGraphicsDropShadowEffect()
    fx.setBlurRadius(blur)
    fx.setColor(QColor(0, 0, 0, alpha))
    fx.setOffset(0, dy)
    widget.setGraphicsEffect(fx)


def clean_code(text: str) -> str:
    """Strip markdown fences and normalise whitespace."""
    text = text.strip()
    # Remove opening fence  ```python  or  ```
    text = re.sub(r'^```[a-zA-Z]*\s*\n?', '', text)
    # Remove closing fence
    text = re.sub(r'\n?```\s*$', '', text)
    return text.strip()


# ---------------------------------------------------------------------------
# Logger (Qt-thread-safe via signals)
# ---------------------------------------------------------------------------
class Logger(QObject):
    log_signal = pyqtSignal(str, str)

    def info(self, msg):    self.log_signal.emit(f"  ○  {msg}", "#6366f1")
    def success(self, msg): self.log_signal.emit(f"  ●  {msg}", "#34d399")
    def error(self, msg):   self.log_signal.emit(f"  ×  {msg}", "#f87171")
    def warn(self, msg):    self.log_signal.emit(f"  ⚠  {msg}", "#fbbf24")


# ---------------------------------------------------------------------------
# Worker thread – calls Groq, executes generated code in sandbox
# ---------------------------------------------------------------------------
class WorkerThread(QThread):
    result_signal = pyqtSignal(str, str, str, dict)
    error_signal  = pyqtSignal(str)

    def __init__(self, client, prompt: str, sandbox: dict):
        super().__init__()
        self.client  = client
        self.prompt  = prompt
        self.sandbox = sandbox          # already a deep copy – safe to mutate

    def run(self):
        try:
            response = self.client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": self.prompt}],
                temperature=0.0,
            )
            raw = response.choices[0].message.content
            code = clean_code(raw)

            buf = io.StringIO()
            with redirect_stdout(buf):
                exec(compile(code, "<ai_code>", "exec"), self.sandbox)  # noqa: S102

            self.result_signal.emit(raw, code, buf.getvalue(), self.sandbox)

        except Exception:
            self.error_signal.emit(traceback.format_exc())


# ---------------------------------------------------------------------------
# Overlay widget shown while AI is running
# ---------------------------------------------------------------------------
class LoadingOverlay(QFrame):
    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, False)
        self.setStyleSheet(
            "background-color: rgba(15,17,23,210); border-radius:16px;"
        )
        lbl = QLabel("⟳  AI is working…", self)
        lbl.setStyleSheet(
            "font-size:20px; font-weight:800; color:#818cf8;"
            " background:transparent; letter-spacing:-0.5px;"
        )
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)

        lay = QVBoxLayout(self)
        lay.addWidget(lbl)
        self.hide()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self.parent():
            self.resize(self.parent().size())


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------
class ExcelArchitectApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.logger = Logger()
        self.logger.log_signal.connect(self._append_log)

        self.current_df: pd.DataFrame | None = None
        self.original_file_path: str | None = None
        self.output_file_path:   str | None = None
        self.df_history: list[pd.DataFrame] = []

        self._worker: WorkerThread | None = None
        self._client = None

        self._build_ui()
        self._init_client()

    # ------------------------------------------------------------------
    # API / client initialisation
    # ------------------------------------------------------------------
    def _init_client(self):
        """Load Groq API key from env or prompt the user."""
        try:
            from dotenv import load_dotenv, set_key
            load_dotenv()
            api_key = os.getenv("GROQ_API_KEY", "").strip()
        except ImportError:
            api_key = os.getenv("GROQ_API_KEY", "").strip()

        if not api_key:
            key, ok = QInputDialog.getText(
                self, "API Key Required",
                "Enter your Groq API Key:",
                QLineEdit.EchoMode.Password,
            )
            if ok and key.strip():
                api_key = key.strip()
                try:
                    from dotenv import set_key
                    set_key(".env", "GROQ_API_KEY", api_key)
                except ImportError:
                    pass
            else:
                self.logger.warn("No API key provided – AI features disabled.")
                return

        try:
            from groq import Groq
            self._client = Groq(api_key=api_key)
            self.logger.success("Groq client ready.")
        except Exception as exc:
            self.logger.error(f"Groq client error: {exc}")

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------
    def _build_ui(self):
        self.setWindowTitle("AI Excel Architect")
        self.resize(1280, 860)
        self.setMinimumSize(1000, 700)
        self.setStyleSheet(STYLE)

        root = QWidget()
        self.setCentralWidget(root)
        vbox = QVBoxLayout(root)
        vbox.setContentsMargins(32, 32, 32, 32)
        vbox.setSpacing(24)

        # ── Header ─────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        tlabel = QLabel("AI Excel Architect")
        tlabel.setObjectName("header")
        badge = QLabel("PRO DATA ENGINE")
        badge.setObjectName("badge")
        badge.setFixedWidth(130)

        title_col = QVBoxLayout()
        title_col.setSpacing(4)
        title_col.addWidget(tlabel)
        title_col.addWidget(badge)

        self.load_btn = QPushButton("⊕  Import Data")
        self.load_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.load_btn.clicked.connect(self._load_file)

        hdr.addLayout(title_col)
        hdr.addStretch()
        hdr.addWidget(self.load_btn)
        vbox.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("background:#2d3148; max-height:1px;")
        vbox.addWidget(sep)

        # ── Main splitter ───────────────────────────────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(1)

        # Left card – data overview
        left = QFrame(); left.setObjectName("card"); drop_shadow(left)
        ll = QVBoxLayout(left); ll.setContentsMargins(22, 22, 22, 22); ll.setSpacing(12)

        ll.addWidget(QLabel(
            "DATA OVERVIEW",
            styleSheet="font-size:11px; font-weight:700; color:#4a5380; letter-spacing:2px;"
        ))

        self.info_label = QLabel("No file loaded yet.")
        self.info_label.setWordWrap(True)
        self.info_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.info_label.setStyleSheet(
            "font-size:13px; color:#64748b; line-height:1.6;"
        )
        self.info_label.setTextFormat(Qt.TextFormat.RichText)

        self.open_btn = QPushButton("Open Result  ↗")
        self.open_btn.setVisible(False)
        self.open_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.open_btn.clicked.connect(self._open_file_external)

        ll.addWidget(self.info_label)
        ll.addStretch()
        ll.addWidget(self.open_btn)

        # Right card – command area
        self._right_card = QFrame(); self._right_card.setObjectName("card")
        drop_shadow(self._right_card)
        rl = QVBoxLayout(self._right_card); rl.setContentsMargins(22, 22, 22, 22); rl.setSpacing(12)

        # Toolbar row
        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel(
            "AI INSTRUCTIONS",
            styleSheet="font-size:11px; font-weight:700; color:#4a5380; letter-spacing:2px;"
        ))
        toolbar.addStretch()
        toolbar.addWidget(QLabel("Recent:", styleSheet="font-size:12px; color:#4a5380;"))
        self._history_cb = QComboBox()
        self._history_cb.setMinimumWidth(210)
        self._history_cb.setPlaceholderText("— history —")
        self._history_cb.activated.connect(self._load_from_history)
        toolbar.addWidget(self._history_cb)
        rl.addLayout(toolbar)

        self._instruction = QTextEdit()
        self._instruction.setPlaceholderText(
            "Describe what you want (e.g. 'Format headers dark blue, freeze row 1, auto-fit columns')…"
        )
        self._instruction.setMinimumHeight(120)
        rl.addWidget(self._instruction)

        # Options row
        opts = QHBoxLayout()
        self._auto_open_cb = QCheckBox("Auto-open result"); self._auto_open_cb.setChecked(True)
        self._show_code_cb = QCheckBox("Show generated code")
        opts.addWidget(self._auto_open_cb)
        opts.addWidget(self._show_code_cb)
        opts.addStretch()
        rl.addLayout(opts)

        # Buttons
        btns = QHBoxLayout(); btns.setSpacing(10)
        self._undo_btn = QPushButton("↩  Undo")
        self._undo_btn.setObjectName("undoBtn")
        self._undo_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._undo_btn.setEnabled(False)
        self._undo_btn.clicked.connect(self._undo)

        self._run_btn = QPushButton("Architect Solution  →")
        self._run_btn.setObjectName("actionButton")
        self._run_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._run_btn.clicked.connect(self._run)

        btns.addWidget(self._undo_btn)
        btns.addWidget(self._run_btn, 1)
        rl.addLayout(btns)

        # Overlay (parented to right card so it covers only that area)
        self._overlay = LoadingOverlay(self._right_card)

        splitter.addWidget(left)
        splitter.addWidget(self._right_card)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        vbox.addWidget(splitter, 1)

        # ── Console footer ──────────────────────────────────────────────
        footer = QFrame(); footer.setObjectName("card"); drop_shadow(footer, blur=16, alpha=40, dy=4)
        fl = QVBoxLayout(footer); fl.setContentsMargins(16, 12, 16, 12); fl.setSpacing(8)

        ctb = QHBoxLayout()
        ctb.addWidget(QLabel(
            "EXECUTION LOG",
            styleSheet="font-size:11px; font-weight:700; color:#4a5380; letter-spacing:2px;"
        ))
        ctb.addStretch()
        clr = QPushButton("Clear"); clr.setStyleSheet("padding:4px 12px; font-size:12px;")
        clr.clicked.connect(lambda: self._console.clear())
        ctb.addWidget(clr)
        fl.addLayout(ctb)

        self._console = QTextEdit()
        self._console.setReadOnly(True)
        self._console.setMinimumHeight(160)
        self._console.setStyleSheet(
            "border:none; background:transparent; font-family:'Cascadia Code','Consolas',monospace; font-size:13px;"
        )
        fl.addWidget(self._console)
        vbox.addWidget(footer)

        self.logger.info("System ready. Import an Excel file to begin.")

    # ------------------------------------------------------------------
    # Overlay helpers
    # ------------------------------------------------------------------
    def _show_overlay(self):
        self._overlay.resize(self._right_card.size())
        self._overlay.show()
        self._overlay.raise_()

    def _hide_overlay(self):
        self._overlay.hide()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._overlay.resize(self._right_card.size())

    # ------------------------------------------------------------------
    # File I/O
    # ------------------------------------------------------------------
    def _load_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Excel File", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if not path:
            return
        try:
            df = pd.read_excel(path)
            self.current_df = df
            self.original_file_path = path
            p = Path(path)
            self.output_file_path = str(p.parent / f"{p.stem}_Updated{p.suffix}")
            self.df_history.clear()
            self._undo_btn.setEnabled(False)
            self._refresh_overview()
            self.open_btn.setVisible(True)
            self.logger.success(f"Loaded '{p.name}' → output: '{p.stem}_Updated{p.suffix}'")
        except Exception as exc:
            self.logger.error(f"Import failed: {exc}")

    def _open_file_external(self):
        target = (
            self.output_file_path
            if self.output_file_path and os.path.exists(self.output_file_path)
            else self.original_file_path
        )
        if not target:
            return
        try:
            if sys.platform == "win32":
                os.startfile(target)  # noqa: S606
            elif sys.platform == "darwin":
                subprocess.call(["open", target])
            else:
                subprocess.call(["xdg-open", target])
        except Exception as exc:
            self.logger.error(f"Could not open file: {exc}")

    # ------------------------------------------------------------------
    # Undo
    # ------------------------------------------------------------------
    def _undo(self):
        if not self.df_history:
            return
        self.current_df = self.df_history.pop()
        try:
            self.current_df.to_excel(self.output_file_path, index=False)
        except Exception as exc:
            self.logger.error(f"Undo save failed: {exc}")
        self._refresh_overview()
        self._undo_btn.setEnabled(bool(self.df_history))
        self.logger.info("Reverted to previous state.")

    # ------------------------------------------------------------------
    # Overview panel
    # ------------------------------------------------------------------
    def _refresh_overview(self):
        df = self.current_df
        if df is None:
            return
        r, c = df.shape
        cols   = df.columns.tolist()
        dtypes = df.dtypes.astype(str).tolist()
        nulls  = df.isnull().sum().tolist()
        uniq   = df.nunique().tolist()

        rows_html = ""
        for i in range(min(c, 14)):
            bg = "#13161f" if i % 2 == 0 else "#1a1d27"
            rows_html += (
                f"<tr style='background:{bg};'>"
                f"<td style='padding:4px 6px; color:#e2e8f0;'>{cols[i]}</td>"
                f"<td style='padding:4px 6px; color:#818cf8;'>{dtypes[i]}</td>"
                f"<td style='padding:4px 6px; color:#f87171;'>{nulls[i]}</td>"
                f"<td style='padding:4px 6px; color:#34d399;'>{uniq[i]}</td>"
                f"</tr>"
            )
        extra = f"<tr><td colspan='4' style='color:#4a5380; font-size:12px; padding:4px 6px;'>+ {c-14} more columns</td></tr>" if c > 14 else ""

        html = (
            f"<div style='margin-bottom:10px; color:#94a3b8;'>"
            f"<b style='color:#e2e8f0;'>{os.path.basename(self.output_file_path)}</b><br>"
            f"{r:,} rows × {c} columns"
            f"</div>"
            f"<table width='100%' style='border-collapse:collapse; font-size:12px;'>"
            f"<tr style='color:#4a5380; font-weight:700;'>"
            f"<td style='padding:4px 6px;'>Column</td>"
            f"<td style='padding:4px 6px;'>Type</td>"
            f"<td style='padding:4px 6px;'>Nulls</td>"
            f"<td style='padding:4px 6px;'>Unique</td>"
            f"</tr>"
            f"{rows_html}{extra}</table>"
        )
        self.info_label.setText(html)

    # ------------------------------------------------------------------
    # History combo
    # ------------------------------------------------------------------
    def _add_to_history(self, cmd: str):
        # Block signals to prevent activated() firing on insertItem
        self._history_cb.blockSignals(True)
        # Remove duplicate if present
        idx = self._history_cb.findText(cmd)
        if idx != -1:
            self._history_cb.removeItem(idx)
        self._history_cb.insertItem(0, cmd)
        if self._history_cb.count() > 15:
            self._history_cb.removeItem(15)
        self._history_cb.setCurrentIndex(0)
        self._history_cb.blockSignals(False)

    def _load_from_history(self, index: int):
        text = self._history_cb.itemText(index)
        if text:
            self._instruction.setPlainText(text)

    # ------------------------------------------------------------------
    # Core: build prompt + run worker
    # ------------------------------------------------------------------
    def _run(self):
        if self.current_df is None:
            self.logger.error("No data loaded. Import an Excel file first.")
            return
        if self._client is None:
            self.logger.error("No API client. Check your Groq API key.")
            return

        cmd = self._instruction.toPlainText().strip()
        if not cmd:
            return

        # Save state for undo
        self.df_history.append(self.current_df.copy())

        # The fix: the sandbox should always load from the existing output file when it exists
        import os
        if self.output_file_path and os.path.exists(self.output_file_path):
            try:
                self.current_df = pd.read_excel(self.output_file_path)
            except Exception as e:
                self.logger.warn(f"Failed to load existing output file for df sync: {e}")

        self._show_overlay()
        self._run_btn.setText("Working…")
        self._run_btn.setEnabled(False)
        self._undo_btn.setEnabled(False)

        # Build context strings
        df = self.current_df
        row_count, col_count = df.shape
        cols_list   = df.columns.tolist()
        sample_str  = df.head(8).to_string(index=False)
        null_str    = df.isnull().sum().to_string()
        unique_str  = df.nunique().to_string()
        dtype_str   = df.dtypes.to_string()

        # NOTE: cols[:2] is rendered now so the prompt is a plain string
        first_col  = cols_list[0] if cols_list else "col0"
        second_col = cols_list[1] if len(cols_list) > 1 else first_col
        two_cols   = str(cols_list[:2])

        prompt = f"""You are a Python code-only machine. Output raw, executable Python code with NO explanation, NO markdown fences, NO comments.

VARIABLES ALREADY IN SCOPE (never redefine or import these):
  df               → pandas DataFrame with the loaded Excel data
  pd               → pandas
  np               → numpy
  os               → os module
  openpyxl         → openpyxl module
  load_workbook    → openpyxl.load_workbook
  output_file_path → string path where the output Excel file must be saved
  PatternFill, Font, Alignment, Border, Side  → from openpyxl.styles
  PieChart, BarChart, LineChart, Reference    → from openpyxl.chart

CURRENT DATASET:
  Rows    : {row_count}
  Columns : {col_count}
  Names   : {cols_list}
  Types   :
{dtype_str}
  Nulls per column:
{null_str}
  Unique per column:
{unique_str}
  First 8 rows:
{sample_str}

═══ FEW-SHOT EXAMPLES ═══

COMMAND: "format the headers with a dark blue background and white bold text, freeze the top row, auto-fit column widths"
CODE:
import os
if os.path.exists(output_file_path):
    wb = load_workbook(output_file_path)
else:
    df.to_excel(output_file_path, index=False)
    wb = load_workbook(output_file_path)
ws = wb.active
for cell in ws[1]:
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws.freeze_panes = 'A2'
for col in ws.columns:
    max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4
wb.save(output_file_path)

COMMAND: "add alternating row colors, light blue and white"
CODE:
import os
if os.path.exists(output_file_path):
    wb = load_workbook(output_file_path)
else:
    df.to_excel(output_file_path, index=False)
    wb = load_workbook(output_file_path)
ws = wb.active
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    color = 'DCE6F1' if i % 2 == 0 else 'FFFFFF'
    for cell in row:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
wb.save(output_file_path)

COMMAND: "remove duplicate rows"
CODE:
before = len(df)
df = df.drop_duplicates(subset={two_cols}, keep='first').reset_index(drop=True)
print(f"Removed {{before - len(df)}} duplicates.")
df.to_excel(output_file_path, index=False)
import os
if os.path.exists(output_file_path):
    wb = load_workbook(output_file_path)
else:
    wb = load_workbook(output_file_path)
ws = wb.active
for cell in ws[1]:
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF')
wb.save(output_file_path)

COMMAND: "add a bar chart showing total values per category"
CODE:
summary_df = df.groupby('{first_col}', as_index=False)['{second_col}'].sum()
summary_df.to_excel(output_file_path, index=False)
import os
if os.path.exists(output_file_path):
    wb = load_workbook(output_file_path)
else:
    wb = load_workbook(output_file_path)
ws = wb.active
for cell in ws[1]:
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF')
chart = BarChart()
data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.title = 'Summary'
ws.add_chart(chart, 'D2')
wb.save(output_file_path)

COMMAND: "fill empty cells in the first column with Pending and highlight them yellow"
CODE:
import os
if os.path.exists(output_file_path):
    wb = load_workbook(output_file_path)
else:
    df.to_excel(output_file_path, index=False)
    wb = load_workbook(output_file_path)
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = 'Pending'
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
wb.save(output_file_path)

COMMAND: "format numeric columns with comma separators, highlight values above 10000 in green"
CODE:
import os
if os.path.exists(output_file_path):
    wb = load_workbook(output_file_path)
else:
    df.to_excel(output_file_path, index=False)
    wb = load_workbook(output_file_path)
ws = wb.active
headers = [cell.value for cell in ws[1]]
numeric_cols = df.select_dtypes(include='number').columns.tolist()
numeric_idx = set(headers.index(c) + 1 for c in numeric_cols if c in headers)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        if cell.column in numeric_idx and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
            if cell.value > 10000:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                cell.font = Font(color='276221')
wb.save(output_file_path)

═══ ABSOLUTE RULES ═══
1. NEVER use Fill() — always PatternFill with start_color, end_color, fill_type='solid'
2. Color strings must NOT have a # prefix — wrong: '#1F4E79'  correct: '1F4E79'
3. NEVER call pd.read_excel() — df is already in scope
4. NEVER call openpyxl.Workbook() — always load_workbook(output_file_path) FIRST if it exists, instead of always calling df.to_excel fresh.
5. NEVER drop_duplicates() without a subset argument
6. NEVER drop columns or rows the user did not mention
7. The final line MUST be wb.save(output_file_path) or df.to_excel(output_file_path, index=False)
8. Use headers list to look up column indices — never hardcode numbers
9. Do not output anything except the raw Python code

NOW COMPLETE THIS TASK:
USER COMMAND: "{cmd}"
"""

        sandbox = {
            "df":               self.current_df.copy(),
            "pd":               pd,
            "np":               np,
            "os":               os,
            "plt":              plt,
            "openpyxl":         openpyxl,
            "load_workbook":    load_workbook,
            "input_file_path":  self.original_file_path,
            "output_file_path": self.output_file_path,
            "PieChart":         PieChart,
            "BarChart":         BarChart,
            "LineChart":        LineChart,
            "Reference":        Reference,
            "PatternFill":      PatternFill,
            "Font":             Font,
            "Alignment":        Alignment,
            "Border":           Border,
            "Side":             Side,
        }

        self._worker = WorkerThread(self._client, prompt, sandbox)
        self._worker.result_signal.connect(self._on_success)
        self._worker.error_signal.connect(self._on_error)
        self._worker.finished.connect(lambda: setattr(self, '_worker', None))
        self._worker.start()

    # ------------------------------------------------------------------
    # Worker callbacks
    # ------------------------------------------------------------------
    def _on_success(self, raw: str, code: str, output: str, sandbox: dict):
        self._hide_overlay()
        self._reset_run_btn()

        cmd = self._instruction.toPlainText().strip()
        self._add_to_history(cmd)

        if self._show_code_cb.isChecked():
            self._log_raw(f"\n── Generated Code ──\n{code}\n", "#7c3aed")

        if output.strip():
            self._log_raw(f"\n── Script Output ──\n{output.strip()}\n", "#64748b")

        # Update dataframe – prefer summary_df if the AI created one
        if "summary_df" in sandbox and isinstance(sandbox["summary_df"], pd.DataFrame):
            self.current_df = sandbox["summary_df"]
        else:
            self.current_df = sandbox["df"]

        # Ensure file was saved (guard for pure df-mutation code without excel ops)
        if "wb.save" not in code and "to_excel" not in code:
            try:
                self.current_df.to_excel(self.output_file_path, index=False)
            except Exception as exc:
                self.logger.error(f"Auto-save failed: {exc}")

        self.logger.success("Solution applied successfully.")
        self._refresh_overview()
        self._undo_btn.setEnabled(True)

        if self._auto_open_cb.isChecked():
            self._open_file_external()

    def _on_error(self, trace: str):
        self._hide_overlay()
        self._run_btn.setText("Failed — Retry →")
        self._run_btn.setStyleSheet(
            "QPushButton#actionButton { background-color:#7f1d1d; color:#fca5a5; border:none; }"
        )
        self._run_btn.setEnabled(True)
        QTimer.singleShot(4000, self._reset_run_btn)

        # Rollback the history entry we optimistically pushed
        if self.df_history:
            self.df_history.pop()
        self._undo_btn.setEnabled(bool(self.df_history))

        self.logger.error("Execution failed — see log below.")
        self._log_raw(f"\n{trace}\n", "#f87171")

    def _reset_run_btn(self):
        self._run_btn.setText("Architect Solution  →")
        self._run_btn.setStyleSheet("")   # returns to QSS objectName styling
        self._run_btn.setEnabled(True)

    # ------------------------------------------------------------------
    # Console helpers
    # ------------------------------------------------------------------
    def _append_log(self, text: str, color: str):
        self._console.moveCursor(QTextCursor.MoveOperation.End)
        self._console.setTextColor(QColor(color))
        self._console.insertPlainText(text + "\n")
        self._console.ensureCursorVisible()

    def _log_raw(self, text: str, color: str):
        self._console.moveCursor(QTextCursor.MoveOperation.End)
        self._console.setTextColor(QColor(color))
        self._console.insertPlainText(text)
        self._console.ensureCursorVisible()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("AI Excel Architect")
    win = ExcelArchitectApp()
    win.show()
    sys.exit(app.exec())